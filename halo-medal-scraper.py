#! python3

import bs4
import openpyxl
import os
import requests
import time
import urllib

game_list = [
    {
        'game': 'halo3-campaign',
        'url': 'https://www.halopedia.org/List_of_Halo_3_and_Halo_3:_ODST_Medals',
        'i': 3
    },
    {
        'game': 'halo3-gmp',
        'url': 'https://www.halopedia.org/List_of_Halo_3_and_Halo_3:_ODST_Medals',
        'i': 0
    },
    {
        'game': 'halo3-omp',
        'url': 'https://www.halopedia.org/List_of_Halo_3_and_Halo_3:_ODST_Medals',
        'i': 1
    },
    {
        'game': 'odst',
        'url': 'https://www.halopedia.org/List_of_Halo_3_and_Halo_3:_ODST_Medals',
        'i': 2
    },
    {
        'game': 'reach',
        'url': 'https://www.halopedia.org/List_of_Halo:_Reach_Medals',
        'i': 0
    },
    {
        'game': 'halo4',
        'url': 'https://www.halopedia.org/List_of_Halo_4_Medals',
        'i': 0
    },
    {
        'game': 'spartan-assault',
        'url': 'https://www.halopedia.org/List_of_Halo:_Spartan_Assault_Medals',
        'i': 0
    },
    {
        'game': 'spartan-strike',
        'url': 'https://www.halopedia.org/List_of_Halo:_Spartan_Strike_Medals',
        'i': 0
    }
]


def save_file(wb):
    wb.save('medal-list.xlsx')
    print('File saved.')


def make_folder(game_name):
    dir = os.path.join('.', 'medals', str(game_name))
    if not os.path.exists(dir):
        os.mkdir(dir)
    return dir


def make_sheet(game_name, wb):
    sheets = wb.sheetnames
    if game_name not in sheets:
        sheet = wb.create_sheet(game_name)
        sheet['A1'] = 'file name'
        sheet['B1'] = 'medal name'
        sheet['C1'] = 'requirement'
        save_file(wb)


def get_soup(link):
    res = requests.get(link)
    res.raise_for_status()
    return bs4.BeautifulSoup(res.text, 'html.parser')


def get_table(game, res):
    elems = res.select('#mw-content-text > table table tbody')
    return elems[game['i']]


def get_medals(table):
    rows = table.select('tr')
    medals = []
    for row in rows:
        if row.find('td'):
            medals.append(row)
    return medals


def log_medals(game, medals, wb):
    print('Logging medals for ' + game['game'] + '...')
    dir = make_folder(game['game'])
    make_sheet(game['game'], wb)
    for medal in medals:
        try:
            file_name = save_medal_pic(dir, medal.select('a')[0])
        except IndexError:
            file_name = 'No image.'
        try:
            log_medal_info(
                game['game'],
                file_name,
                medal.select('td')[1].text.strip(),
                medal.select('td')[2].text.strip(),
                wb
            )
        except IndexError:
            continue
    mark_complete(game['game'], wb)


def save_medal_pic(dir, medal):
    file_name = medal.get('href')[6:]
    base = 'https://www.halopedia.org'
    soup = get_soup(urllib.parse.urljoin(base, medal.get('href')))
    img_link = soup.select('#file > a > img')[0].get('src')
    res = requests.get(urllib.parse.urljoin(base, img_link))
    res.raise_for_status()
    img = open(os.path.join(dir, file_name), 'wb')
    img.write(res.content)
    img.close()
    return file_name


def log_medal_info(sheet_name, file_name, medal_name, requirement, wb):
    sheet = wb[sheet_name]
    row = (file_name, medal_name, requirement)
    sheet.append(row)


def mark_complete(game_name, wb):
    sheet = wb['checklist']
    sheet.append([game_name])
    save_file(wb)
    print('All medals logged for ' + game_name + '.')


def scrape_medals(games):
    wb = openpyxl.load_workbook('medal-list.xlsx')
    for game in games:
        table = get_table(game, get_soup(game['url']))
        medals = get_medals(table)
        log_medals(game, medals, wb)
    save_file(wb)
    wb.close()


if __name__ == "__main__":
    start_time = time.time()
    print('Beginning medal scraping.')
    scrape_medals(game_list)
    print('All medals scraped in %s seconds.' % (round(time.time() - start_time)))
